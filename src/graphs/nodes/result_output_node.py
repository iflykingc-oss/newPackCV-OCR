# -*- coding: utf-8 -*-
"""
结果输出节点
将结构化数据格式化输出，支持JSON/Excel/PDF格式
"""

import os
import json
import time
import logging
from typing import Dict, Any, Optional
from io import BytesIO

from langchain_core.runnables import RunnableConfig
from langgraph.runtime import Runtime
from coze_coding_utils.runtime_ctx.context import Context
from graphs.state import ResultOutputInput, ResultOutputOutput
from coze_coding_dev_sdk.s3 import S3SyncStorage

logger = logging.getLogger(__name__)


def _get_s3_storage() -> S3SyncStorage:
    """获取S3对象存储客户端"""
    return S3SyncStorage(
        endpoint_url=os.getenv("COZE_BUCKET_ENDPOINT_URL"),
        access_key="",
        secret_key="",
        bucket_name=os.getenv("COZE_BUCKET_NAME"),
        region="cn-beijing",
    )


def export_to_json(data: Dict[str, Any]) -> str:
    """导出为JSON格式"""
    return json.dumps(data, ensure_ascii=False, indent=2)


def export_to_excel(data: Dict[str, Any]) -> Optional[bytes]:
    """导出为Excel格式"""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            return None
        ws.title = "OCR识别结果"

        # 写入表头
        headers = ["字段", "值"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # 写入数据
        row = 2
        for key, value in data.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value) if value is not None else "")
            row += 1

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    except Exception as e:
        logger.warning(f"Excel导出失败: {e}")
        return None


def export_to_pdf(data: Dict[str, Any]) -> Optional[bytes]:
    """导出为PDF格式"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas

        output = BytesIO()
        c = pdf_canvas.Canvas(output, pagesize=A4)
        width, height = A4

        c.setFont("Helvetica", 16)
        c.drawString(50, height - 50, "OCR Recognition Result")

        c.setFont("Helvetica", 10)
        y_position = height - 80

        for key, value in data.items():
            if y_position < 50:
                c.showPage()
                y_position = height - 50

            line = f"{key}: {value}"
            c.drawString(50, y_position, line)
            y_position -= 20

        c.save()
        return output.getvalue()
    except Exception as e:
        logger.warning(f"PDF导出失败（reportlab可能未安装）: {e}")
        return None


def result_output_node(
    state: ResultOutputInput,
    config: RunnableConfig,
    runtime: Runtime[Context]
) -> ResultOutputOutput:
    """
    title: 结果输出
    desc: 将结构化数据格式化输出，支持JSON/Excel/PDF格式，并上传到对象存储
    integrations: 对象存储
    """
    ctx = runtime.context

    # 组装输出数据
    output_data: Dict[str, Any] = {
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "structured_data": state.structured_data or {},
    }

    # V5.4 商业化统一结构：附加 category_info / warnings / ext_info
    if state.category_info:
        output_data["category_info"] = state.category_info
    if state.warnings:
        output_data["warnings"] = state.warnings
    if state.ext_info:
        output_data["ext_info"] = state.ext_info

    if state.corrected_text:
        output_data["corrected_text"] = state.corrected_text
    elif state.corrected_result:
        output_data["corrected_text"] = state.corrected_result
    if state.raw_text:
        output_data["raw_text"] = state.raw_text
    elif state.ocr_raw_result:
        output_data["raw_text"] = state.ocr_raw_result
    elif state.ocr_result:
        output_data["raw_text"] = state.ocr_result
    if state.answer:
        output_data["answer"] = state.answer
    elif state.qa_answer:
        output_data["answer"] = state.qa_answer

    export_format = state.export_format or "json"
    export_file_url = ""

    try:
        storage = _get_s3_storage()

        if export_format == "json":
            json_content = export_to_json(output_data)
            file_bytes = json_content.encode('utf-8')
            file_name = f"ocr_output/result_{int(time.time())}.json"
            content_type = 'application/json'

            key = storage.upload_file(
                file_content=file_bytes,
                file_name=file_name,
                content_type=content_type
            )
            export_file_url = storage.generate_presigned_url(key=key, expire_time=3600)

        elif export_format == "excel":
            excel_bytes = export_to_excel(output_data)
            if excel_bytes:
                file_name = f"ocr_output/result_{int(time.time())}.xlsx"
                key = storage.upload_file(
                    file_content=excel_bytes,
                    file_name=file_name,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                export_file_url = storage.generate_presigned_url(key=key, expire_time=3600)
            else:
                # Excel导出失败，降级到JSON
                json_content = export_to_json(output_data)
                file_bytes = json_content.encode('utf-8')
                key = storage.upload_file(file_content=file_bytes, file_name=f"ocr_output/result_{int(time.time())}.json", content_type='application/json')
                export_file_url = storage.generate_presigned_url(key=key, expire_time=3600)

        elif export_format == "pdf":
            pdf_bytes = export_to_pdf(output_data)
            if pdf_bytes:
                file_name = f"ocr_output/result_{int(time.time())}.pdf"
                key = storage.upload_file(
                    file_content=pdf_bytes,
                    file_name=file_name,
                    content_type='application/pdf'
                )
                export_file_url = storage.generate_presigned_url(key=key, expire_time=3600)
            else:
                # PDF导出失败，降级到JSON
                json_content = export_to_json(output_data)
                file_bytes = json_content.encode('utf-8')
                key = storage.upload_file(file_content=file_bytes, file_name=f"ocr_output/result_{int(time.time())}.json", content_type='application/json')
                export_file_url = storage.generate_presigned_url(key=key, expire_time=3600)

        logger.info(f"结果已上传S3，格式: {export_format}")

    except Exception as e:
        logger.warning(f"S3上传失败: {e}")

    # 平台推送（可选）
    platform_push_result: Dict[str, Any] = {}
    if state.platform and state.platform != "none":
        try:
            platform_push_result = _push_to_platform(state.platform, output_data)
        except Exception as e:
            logger.warning(f"平台推送失败: {e}")
            platform_push_result = {"error": str(e)}

    return ResultOutputOutput(
        final_result=output_data,
        export_file_url=export_file_url,
        platform_push_result=platform_push_result
    )


def _push_to_platform(platform: str, data: Dict[str, Any]) -> Dict[str, Any]:
    """推送到指定平台"""
    result: Dict[str, Any] = {"platform": platform, "status": "skipped"}
    if platform in ("feishu", "wechat", "dingtalk"):
        logger.info(f"推送通知到 {platform}（功能待集成）")
        result = {"platform": platform, "status": "pending_integration"}
    else:
        logger.info(f"未知平台: {platform}，跳过推送")
    return result
